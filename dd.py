import re
import requests
import urllib.parse
import pandas as pd
import tempfile
import json
import threading
from concurrent.futures import ThreadPoolExecutor, as_completed
from bs4 import BeautifulSoup
from flask import Flask, request, jsonify, render_template, send_file
from playwright.sync_api import sync_playwright
from datetime import datetime


# ------------------------------
# Utils
# ------------------------------
def format_price(n):
    if not n:
        return "غير متوفر"
    try:
        return f"{float(n):,.0f} رس"
    except:
        return "غير متوفر"


def launch_playwright(url, wait_key=None, extra_headers=None):
    """فتح متصفح جديد لجلب صفحة - بدون إعادة محاولة"""
    p = sync_playwright().start()
    try:
        browser = p.chromium.launch(
            headless=True,
            args=['--no-sandbox', '--disable-setuid-sandbox',
                  '--disable-dev-shm-usage', '--disable-gpu',
                  '--single-process', '--disable-blink-features=AutomationControlled']
        )
        ctx_args = dict(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
            viewport={"width": 1366, "height": 768}
        )
        if extra_headers:
            ctx_args["extra_http_headers"] = extra_headers
            ctx_args["locale"] = "ar-SA"

        ctx = browser.new_context(**ctx_args)
        page = ctx.new_page()
        page.add_init_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")

        # تقليل وقت التحميل
        response = page.goto(url, timeout=30000, wait_until='domcontentloaded')
        if not response or response.status >= 400:
            browser.close()
            p.stop()
            return None

        # انتظار المحددات بمهلة أقل
        if wait_key == "mahally":
            page.wait_for_selector("a.styles_productCard__name__pakbB", timeout=8000, state='attached')
        elif wait_key == "almanea":
            page.wait_for_selector("div.relative.h-full.overflow-hidden.bg-white.rounded-xl", timeout=8000)
        elif wait_key == "closebuy":
            page.wait_for_selector("div.product.product-1", timeout=8000)
        elif wait_key == "extra":
            # تم إلغاء النقر على الكوكيز لتوفير الوقت
            page.wait_for_timeout(1500)
        elif wait_key == "tamkeen":
            page.wait_for_timeout(3000)
        else:
            page.wait_for_timeout(1500)

        html = page.content()
        browser.close()
        p.stop()
        return html
    except Exception as e:
        try:
            p.stop()
        except:
            pass
        return None  # لا إعادة محاولة


# ------------------------------
# Base Store Class
# ------------------------------
class StoreBase:
    def __init__(self, key, display_name, base_url, query_fmt):
        self.key = key
        self.display_name = display_name
        self.base_url = base_url
        self.query_fmt = query_fmt

    def build_search_url(self, query):
        return self.base_url + self.query_fmt.format(query=urllib.parse.quote(query))

    def fetch(self, url):
        # محاولة واحدة فقط
        return launch_playwright(url, wait_key=self.key)

    def fetch_requests(self, url):
        try:
            r = requests.get(url, timeout=15, headers={"User-Agent": "Mozilla/5.0"})
            return r.text if r.status_code == 200 else None
        except:
            return None

    def clean_text(self, txt):
        return re.sub(r"\s+", " ", txt).strip() if txt else ""

    def extract_price(self, txt):
        if not txt:
            return None
        txt = txt.replace(",", "").replace("٫", ".")
        m = re.search(r"(\d+(?:\.\d+)?)", txt)
        return float(m.group(1)) if m else None

    def smart_match(self, query, name):
        if not query or not name:
            return True
        q = re.sub(r"[\s\-_]", "", query.lower())
        n = re.sub(r"[\s\-_]", "", name.lower())
        if q in n:
            return True
        q_nums = re.findall(r"\d+", q)
        n_nums = re.findall(r"\d+", n)
        if q_nums and n_nums:
            for qn in q_nums:
                for nn in n_nums:
                    if qn == nn or (len(qn) >= 3 and qn in nn) or (len(nn) >= 3 and nn in qn):
                        return True
            if q_nums and not n_nums:
                return True
        for letters in re.findall(r"[a-z]{2,}", q):
            if letters in n:
                return True
        return False

    def match_score(self, name, query):
        if not name or not query:
            return 50
        n, q = name.lower(), query.lower()
        if q in n:
            return 100
        return 50 + sum(10 for p in q.split() if len(p) > 1 and p in n)

    def sort_results(self, results, query):
        real = [r for r in results if not r.get("_fb")]
        fb = [r for r in results if r.get("_fb")]
        if real:
            return sorted(real, key=lambda x: x["match_score"], reverse=True)
        return fb[:1] if fb else []


# ------------------------------
# Extra Store
# ------------------------------
class ExtraStore(StoreBase):
    def __init__(self):
        super().__init__('extra', '🛒 إكسترا', 'https://www.extra.com', '/ar-sa/search/?text={query}')

    def parse(self, html, query):
        if not html:
            return []
        soup = BeautifulSoup(html, "html.parser")
        cards = soup.find_all("div", class_=lambda x: x and "product-tile-container" in x)
        if not cards:
            cards = soup.find_all("section", class_=lambda x: x and "product-tile-wrapper" in x)
        results = []
        for card in cards:
            try:
                brand_el = card.select_one(".brand-name") or card.select_one("[class*='brand']")
                title_el = (card.select_one(".product-name-data") or
                            card.select_one("[class*='product-name']") or
                            card.select_one("h2") or card.select_one("h3"))
                if not title_el:
                    continue
                brand = brand_el.get_text(strip=True) if brand_el else ""
                title = title_el.get_text(strip=True)
                name = f"{brand} - {title}".strip(" -") if brand else title
                link_el = card.find("a", href=True)
                href = link_el.get("href", "") if link_el else ""
                link = href if href.startswith("http") else self.base_url + href
                price_el = card.find("strong")
                current_price = self.extract_price(price_el.get_text(strip=True)) if price_el else None
                old_el = card.find("span", class_=lambda x: x and "striked-off" in x)
                old_price = self.extract_price(old_el.get_text(strip=True)) if old_el else None
                discount = ""
                if current_price and old_price and old_price > current_price:
                    discount = f"{round(((old_price - current_price) / old_price) * 100)}%"
                is_match = self.smart_match(query, name)
                results.append({
                    "name": name,
                    "current_price": format_price(current_price) if current_price else "غير متوفر",
                    "original_price": format_price(old_price) if old_price else "",
                    "discount": discount,
                    "store": self.display_name,
                    "link": link,
                    "match_score": self.match_score(name, query) if is_match else 5,
                    "_fb": not is_match
                })
            except Exception:
                pass
        return self.sort_results(results, query)


# ------------------------------
# Al Manea Store
# ------------------------------
class AlmaneaStore(StoreBase):
    def __init__(self):
        super().__init__("almanea", "🏪 المنيع", "https://www.almanea.sa", "/search?q={query}")

    def parse(self, html, query):
        if not html:
            return []
        soup = BeautifulSoup(html, "html.parser")
        results = []
        cards = soup.select("div.relative.h-full.overflow-hidden.bg-white.rounded-xl")
        for card in cards:
            try:
                name_el = card.select_one("a.cursor-pointer span")
                if not name_el:
                    continue
                name = name_el.get_text(strip=True)
                if not self.smart_match(query, name):
                    continue
                link_el = card.select_one("a.cursor-pointer")
                link = self.base_url + link_el["href"] if link_el and link_el.has_attr("href") else ""
                new_price_el = card.select_one("p.text-red span")
                new_price = self.extract_price(new_price_el.get_text(strip=True)) if new_price_el else None
                old_price_el = card.select_one("p.line-through span")
                old_price = self.extract_price(old_price_el.get_text(strip=True)) if old_price_el else None
                discount = ""
                if new_price and old_price and old_price > new_price:
                    discount = f"{round(((old_price - new_price) / old_price) * 100)}%"
                results.append({
                    "name": name, "current_price": format_price(new_price) if new_price else "غير متوفر",
                    "original_price": format_price(old_price) if old_price else "",
                    "discount": discount, "store": self.display_name, "link": link,
                    "match_score": self.match_score(name, query)
                })
            except Exception:
                pass
        return sorted(results, key=lambda x: x["match_score"], reverse=True)


# ------------------------------
# CloseBuy Store
# ------------------------------
class CloseBuyStore(StoreBase):
    def __init__(self):
        super().__init__("closebuy", "🛍️ كلوز باي", "https://closebuy.sa", "/products?q={query}")

    def parse(self, html, query):
        if not html:
            return []
        soup = BeautifulSoup(html, "html.parser")
        results = []
        cards = soup.select("div.product.product-1")
        for card in cards:
            try:
                name_el = card.select_one(".product-bottom .title a")
                if not name_el:
                    continue
                name = self.clean_text(name_el.get_text())
                if not self.smart_match(query, name):
                    continue
                link = name_el.get("href", "")
                if link and not link.startswith("http"):
                    link = self.base_url + link
                new_price_el = card.select_one(".discount-price")
                new_price = self.extract_price(new_price_el.get_text(strip=True)) if new_price_el else None
                old_price_el = card.select_one("del.nondiscount-price")
                old_price = self.extract_price(old_price_el.get_text(strip=True)) if old_price_el else None
                discount = ""
                discount_el = card.select_one(".percentage")
                if discount_el:
                    discount = discount_el.get_text(strip=True)
                elif new_price and old_price and old_price > new_price:
                    discount = f"خصم {round(((old_price - new_price) / old_price) * 100)}%"
                if new_price:
                    results.append({
                        "name": name, "current_price": format_price(new_price),
                        "original_price": format_price(old_price) if old_price else "",
                        "discount": discount, "store": self.display_name, "link": link,
                        "match_score": self.match_score(name, query)
                    })
            except Exception:
                pass
        return sorted(results, key=lambda x: x["match_score"], reverse=True)


# ------------------------------
# Mahally Store
# ------------------------------
class MahallyStore(StoreBase):
    def __init__(self):
        super().__init__("mahally", "🏬 محلي", "https://mahally.com", "/browse/?query={query}")

    def parse(self, html, query):
        if not html:
            return []
        soup = BeautifulSoup(html, "html.parser")
        results = []
        for card in soup.select("div.relative.h-full"):
            try:
                name_el = card.select_one("a.styles_productCard__name__pakbB")
                if not name_el:
                    continue
                name = self.clean_text(name_el.text)
                if not self.smart_match(query, name):
                    continue
                link_el = card.select_one("a[aria-label='Product Details']")
                link = self.base_url + link_el["href"] if link_el and link_el.has_attr("href") else ""
                new_el = card.select_one(".styles_productCard__price__uGOio span")
                old_el = card.select_one(".styles_productCard__salePrice__LreTD span")
                results.append({
                    "name": name,
                    "current_price": format_price(self.extract_price(new_el.text) if new_el else None),
                    "original_price": format_price(
                        self.extract_price(old_el.text) if old_el else None) if old_el else "",
                    "store": self.display_name, "link": link,
                    "match_score": self.match_score(name, query)
                })
            except Exception:
                pass
        return sorted(results, key=lambda x: x["match_score"], reverse=True)


# ------------------------------
# SWSG Store
# ------------------------------
class SwsgStore(StoreBase):
    def __init__(self):
        super().__init__("swsg", "🛒 SWSG الصندوق الأسود", "https://swsg.co", "/ar/search/?q={query}")

    def fetch(self, url):
        return launch_playwright(url, wait_key=self.key, extra_headers={
            "Accept-Language": "ar-SA,ar;q=0.9,en;q=0.8",
            "Referer": "https://swsg.co/ar",
        })

    def parse(self, html, query):
        if not html or len(html) < 5000:
            return []
        soup = BeautifulSoup(html, "html.parser")
        results = []
        cards = soup.select("div.kuProdBottom")
        if not cards:
            cards = soup.find_all("div", class_=lambda x: x and "kuProd" in str(x))
        for card in cards:
            try:
                name_el = card.select_one(".kuName a")
                if not name_el:
                    continue
                name = self.clean_text(name_el.get_text())
                if not self.smart_match(query, name):
                    continue
                link_el = card.find("a", href=True)
                href = link_el.get("href", "") if link_el else ""
                link = href if href.startswith("http") else self.base_url + href
                price_el = card.select_one(".kuSalePrice")
                new_price = self.extract_price(price_el.get_text(strip=True)) if price_el else None
                old_el = card.select_one(".kuOrigPrice")
                old_price = self.extract_price(old_el.get_text(strip=True)) if old_el else None
                discount = ""
                if new_price and old_price and old_price > new_price:
                    discount = f"خصم {round(((old_price - new_price) / old_price) * 100)}%"
                if new_price:
                    results.append({
                        "name": name, "current_price": format_price(new_price),
                        "original_price": format_price(old_price) if old_price else "",
                        "discount": discount, "store": self.display_name, "link": link,
                        "match_score": self.match_score(name, query)
                    })
            except Exception:
                pass
        return sorted(results, key=lambda x: x["match_score"], reverse=True)


# ------------------------------
# AlGhanim Store
# ------------------------------
class AlGhanimStore(StoreBase):
    def __init__(self):
        super().__init__('alghanim', '🔵 الغانم', 'https://alghanim-store.com', '/search?q={query}')

    def parse(self, html, query):
        if not html:
            return []
        soup = BeautifulSoup(html, "html.parser")
        results = []
        q_nums = re.findall(r"\d+", query)
        product_links = list(set(
            a['href'] for a in soup.find_all('a', href=True)
            if ('/product/' in a['href'] or '/products/' in a['href'])
            and (any(n in a.get_text() or n in a['href'] for n in q_nums) if q_nums
                 else query.lower() in a.get_text().lower())
        ))[:3]
        for link in product_links:
            try:
                product_url = link if link.startswith('http') else self.base_url + link
                product_html = launch_playwright(product_url)  # محاولة واحدة فقط
                if not product_html:
                    continue
                psoup = BeautifulSoup(product_html, "html.parser")
                name_el = psoup.find('h1') or psoup.find('meta', {'property': 'og:title'})
                if not name_el:
                    continue
                name = name_el.get('content', '') if name_el.name == 'meta' else name_el.get_text(strip=True)
                if not name or "لا توجد نتائج" in name:
                    continue
                current_price = None
                for pat in [r'السعر الحالي هو:?\s*(\d+)', r'(\d+)\s*\(شامل الضريبة\)', r'(\d+)\s*ريال', r'(\d+)\s*رس']:
                    m = re.search(pat, product_html)
                    if m:
                        current_price = self.extract_price(m.group(1))
                        break
                if current_price and name:
                    results.append({
                        "name": name, "current_price": format_price(current_price),
                        "original_price": "", "discount": "",
                        "store": self.display_name, "link": product_url, "match_score": 100
                    })
            except Exception:
                pass
        return results


# ------------------------------
# Tamkeen Store
# ------------------------------
class TamkeenStore(StoreBase):
    def __init__(self):
        super().__init__("tamkeen", "🟣 تمكين", "https://tamkeenstores.com.sa", "/ar/search?text={query}")

    def parse(self, html, query):
        if not html:
            return []
        soup = BeautifulSoup(html, "html.parser")
        results = []
        cards = soup.find_all("div", class_=lambda x: x and "product_card" in x)
        if not cards:
            cards = soup.find_all("div", class_=lambda x: x and "relative" in x and "h-full" in x)
        for card in cards:
            try:
                name_el = card.find("div", class_=lambda x: x and "line-clamp-2" in x)
                if not name_el:
                    continue
                name = self.clean_text(name_el.get_text())
                if not name or len(name) < 3:
                    continue
                link_el = card.find("a", href=True)
                href = link_el.get("href", "") if link_el else ""
                link = self.base_url + href if href.startswith("/") else href
                price_el = card.find("span", class_=lambda x: x and "font-bold" in x)
                new_price = self.extract_price(price_el.get_text()) if price_el else None
                if not new_price:
                    continue
                old_price_el = card.find("span", class_=lambda x: x and "line-through" in x)
                old_price = self.extract_price(old_price_el.get_text()) if old_price_el else None
                discount_el = card.find("div", class_=lambda x: x and "bg-danger" in x)
                discount = discount_el.get_text(strip=True) if discount_el else ""
                is_match = self.smart_match(query, name)
                results.append({
                    "name": name, "current_price": format_price(new_price),
                    "original_price": format_price(old_price) if old_price else "",
                    "discount": discount, "store": self.display_name, "link": link,
                    "match_score": self.match_score(name, query) if is_match else 5,
                    "_fb": not is_match
                })
            except Exception:
                pass
        return self.sort_results(results, query)


# ------------------------------
# Scraper Engine with Parallel Store Fetching
# ------------------------------
class ScraperEngine:
    def __init__(self, stores):
        self.stores = {s.key: s for s in stores}

    def _search_one(self, key, query):
        """جلب متجر واحد (يُستخدم في الخيوط المتوازية)"""
        if key not in self.stores:
            return key, []
        store = self.stores[key]
        url = store.build_search_url(query)
        try:
            html = store.fetch(url)
            if html:
                results = store.parse(html, query)
                return key, results
            return key, []
        except Exception:
            return key, []

    def search(self, query, store_keys, max_workers=5):
        """بحث متوازي عن المنتج في عدة متاجر"""
        print(f"🔍 '{query}' في {len(store_keys)} متجر (بالتوازي)")
        out = {}
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            future_to_key = {executor.submit(self._search_one, k, query): k for k in store_keys}
            for future in as_completed(future_to_key):
                key, results = future.result()
                out[key] = results
        return out


# ------------------------------
# Flask App
# ------------------------------
app = Flask(__name__)
stores = [ExtraStore(), AlmaneaStore(), MahallyStore(), SwsgStore(), CloseBuyStore(), AlGhanimStore(), TamkeenStore()]
engine = ScraperEngine(stores)

STORE_DISPLAY = {'extra': '🛒 إكسترا', 'almanea': '🏪 المنيع', 'mahally': '🏬 محلي',
                 'closebuy': '🛍️ كلوز باي', 'swsg': '🛒 SWSG', 'alghanim': '🔵 الغانم', 'tamkeen': '🟣 تمكين'}
STORE_COLORS = {'🛒 إكسترا': 'FFC7CE', '🏪 المنيع': 'C6EFCE', '🏬 محلي': 'FFEB9C',
                '🛍️ كلوز باي': 'B7DEE8', '🛒 SWSG': 'E2D1FF', '🔵 الغانم': 'B4C6E7', '🟣 تمكين': 'F9D7E2'}


@app.route("/")
def home():
    return render_template("index.html")


@app.route("/search", methods=["POST"])
def api_search():
    data = request.json
    return jsonify(engine.search(data.get("query"), data.get("stores", [])))


@app.route("/process_excel", methods=["POST"])
def process_excel():
    file = request.files.get("file")
    if not file:
        return "No file", 400

    df = pd.read_excel(file)
    models = df.iloc[:, 0].dropna().astype(str).tolist()
    try:
        store_keys = json.loads(request.form.get("stores", "[]"))
    except:
        store_keys = list(STORE_DISPLAY.keys())

    # بحث متوازي عن الموديلات (حد أقصى 3 موديلات في وقت واحد)
    results_dict = {}
    with ThreadPoolExecutor(max_workers=3) as executor:
        future_to_model = {executor.submit(engine.search, model, store_keys): model for model in models}
        for future in as_completed(future_to_model):
            model = future_to_model[future]
            try:
                search_result = future.result()
                model_row = {'الموديل': model}
                for sk in store_keys:
                    sn = STORE_DISPLAY.get(sk, sk)
                    products = search_result.get(sk, [])
                    if products:
                        p = products[0]
                        model_row[f'{sn} (السعر)'] = p.get('current_price', 'غير متوفر')
                        model_row[f'{sn} (قديم)'] = p.get('original_price', '')
                        model_row[f'{sn} (خصم)'] = p.get('discount', '')
                        model_row[f'{sn} (رابط)'] = p.get('link', '')
                    else:
                        model_row[f'{sn} (السعر)'] = 'غير متوفر'
                        model_row[f'{sn} (قديم)'] = model_row[f'{sn} (خصم)'] = model_row[f'{sn} (رابط)'] = ''
                results_dict[model] = model_row
            except Exception as e:
                print(f"خطأ في معالجة {model}: {e}")

    result_df = pd.DataFrame.from_dict(results_dict, orient='index')
    columns = ['الموديل']
    for sk in store_keys:
        sn = STORE_DISPLAY.get(sk, sk)
        columns.extend([f'{sn} (السعر)', f'{sn} (قديم)', f'{sn} (خصم)', f'{sn} (رابط)'])
    for col in columns:
        if col not in result_df.columns:
            result_df[col] = ''
    result_df = result_df[columns]

    temp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    with pd.ExcelWriter(temp.name, engine='openpyxl') as writer:
        result_df.to_excel(writer, index=False, sheet_name='مقارنة الأسعار')
        from openpyxl.styles import PatternFill, Font
        ws = writer.sheets['مقارنة الأسعار']
        for ci, cn in enumerate(columns, 1):
            cell = ws.cell(row=1, column=ci)
            cell.font = Font(bold=True)
            for sn, color in STORE_COLORS.items():
                if sn in cn:
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    break
        for ri in range(2, ws.max_row + 1):
            for ci, cn in enumerate(columns, 1):
                cell = ws.cell(row=ri, column=ci)
                for sn, color in STORE_COLORS.items():
                    if sn in cn:
                        cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                        break
        for col in ws.columns:
            max_len = max((len(str(c.value)) for c in col if c.value), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    return send_file(temp.name, as_attachment=True,
                     download_name=f"مقارنة_الأسعار_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


if __name__ == "__main__":
    app.run(debug=False, host='0.0.0.0', port=5001)